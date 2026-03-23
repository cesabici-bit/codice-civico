"""Ministero della Giustizia court statistics ingestor.

Data source: https://datiestatistiche.giustizia.it/page/it/flussi-tribunali-ordinari-e-corti-di-appello
Format: Excel .xlsx files (CivileFlussi2014-2024.xlsx and similar)

Metrics per tribunal/year/category:
- Iscritti (incoming / new_cases)
- Definiti (resolved / resolved_cases)
- Pendenti finali (pending / pending_cases)
- Clearance Rate = definiti / iscritti  (computed)
- Disposition Time = (pendenti / definiti) * 365  (computed, in days)
"""

from __future__ import annotations

import io
import logging
from datetime import date
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from codicecivico.config import settings
from codicecivico.ingest.base import BaseIngestor
from codicecivico.ingest.tribunali_seed import TribunalSeed, get_tribunali
from codicecivico.models import CourtStat, Tribunal

logger = logging.getLogger(__name__)

# Known Excel download URLs (verified 2026-03-23)
FLUSSI_CIVILE_URL = (
    "https://datiestatistiche.giustizia.it/resources/"
    "CivileFlussi2014-2024.xlsx"
)

HTTP_TIMEOUT = 120.0  # seconds — Excel files can be large
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
}

# Column name variants (the Excel structure may vary)
_COL_ALIASES: dict[str, list[str]] = {
    "district": ["Distretto", "distretto", "DISTRETTO", "Corte di Appello"],
    "tribunal": [
        "Circondario", "circondario", "CIRCONDARIO",
        "Tribunale", "tribunale", "TRIBUNALE", "Ufficio",
    ],
    "year": ["Anno", "anno", "ANNO", "Periodo"],
    "incoming": [
        "Iscritti", "iscritti", "ISCRITTI",
        "Sopravvenuti", "sopravvenuti", "Nuovi iscritti",
    ],
    "resolved": [
        "Definiti", "definiti", "DEFINITI",
        "Esauriti", "esauriti",
    ],
    "pending": [
        "Pendenti finali", "pendenti finali", "PENDENTI FINALI",
        "Pendenti", "pendenti", "Pendenze finali",
    ],
}


# ---------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------


class CourtRecord:
    """Parsed row from the Excel file."""

    __slots__ = (
        "tribunal_name", "district", "year", "case_category",
        "incoming", "resolved", "pending",
        "clearance_rate", "disposition_time",
    )

    def __init__(
        self,
        tribunal_name: str,
        district: str,
        year: int,
        case_category: str,
        incoming: int,
        resolved: int,
        pending: int,
    ) -> None:
        self.tribunal_name = tribunal_name
        self.district = district
        self.year = year
        self.case_category = case_category
        self.incoming = incoming
        self.resolved = resolved
        self.pending = pending
        self.clearance_rate = compute_clearance_rate(resolved, incoming)
        self.disposition_time = compute_disposition_time(pending, resolved)


# ---------------------------------------------------------------
# Metric computation (pure functions, easily testable)
# ---------------------------------------------------------------


def compute_clearance_rate(resolved: int, incoming: int) -> float | None:
    """Clearance rate = resolved / incoming.

    Returns None if incoming == 0 (avoid division by zero).
    Values > 1.0 mean the tribunal is reducing its backlog.
    """
    if incoming <= 0:
        return None
    return round(resolved / incoming, 4)


def compute_disposition_time(pending: int, resolved: int) -> float | None:
    """Disposition time = (pending / resolved) * 365 days.

    Standard CEPEJ formula. Returns None if resolved == 0.
    """
    if resolved <= 0:
        return None
    return round((pending / resolved) * 365, 2)


# ---------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------


def _resolve_columns(
    header_row: tuple[Any, ...],
) -> dict[str, int]:
    """Map logical column names to 0-based column indices.

    Tries each alias for each logical name against the header row.
    Raises ValueError if a required column is not found.
    """
    header_values = [str(cell).strip() if cell else "" for cell in header_row]
    mapping: dict[str, int] = {}

    for logical, aliases in _COL_ALIASES.items():
        for alias in aliases:
            for idx, header_val in enumerate(header_values):
                if header_val == alias:
                    mapping[logical] = idx
                    break
            if logical in mapping:
                break

    required = {"tribunal", "year", "incoming", "resolved", "pending"}
    missing = required - set(mapping.keys())
    if missing:
        msg = f"Missing required columns: {missing}. Found headers: {header_values}"
        raise ValueError(msg)

    return mapping


def _safe_int(value: Any) -> int:
    """Convert a cell value to int, handling None and float."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = str(value).strip().replace(",", "").replace(".", "")
    if not cleaned or cleaned == "-":
        return 0
    return int(cleaned)


def _safe_year(value: Any) -> int | None:
    """Extract a 4-digit year from a cell value."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        y = int(value)
        if 2000 <= y <= 2100:
            return y
        return None
    s = str(value).strip()
    # Try extracting first 4 digits
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) >= 4:
        y = int(digits[:4])
        if 2000 <= y <= 2100:
            return y
    return None


def parse_excel(
    file_content: bytes,
    *,
    sheet_name: str | None = None,
    case_category: str = "civile",
    header_row_idx: int = 3,
) -> list[CourtRecord]:
    """Parse a Ministero Giustizia Excel file into CourtRecord list.

    Args:
        file_content: Raw bytes of the .xlsx file.
        sheet_name: Worksheet name (None = first/active sheet).
        case_category: Category label for these records.
        header_row_idx: 1-based row number containing column headers.

    Returns:
        List of CourtRecord, one per tribunal/year combination.
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active  # type: ignore[assignment]
    if ws is None:
        msg = "No active worksheet found"
        raise ValueError(msg)

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < header_row_idx:
        msg = f"Excel has only {len(rows)} rows, expected header at row {header_row_idx}"
        raise ValueError(msg)

    header = rows[header_row_idx - 1]
    col_map = _resolve_columns(header)

    records: list[CourtRecord] = []
    for row in rows[header_row_idx:]:
        trib_idx = col_map.get("tribunal")
        tribunal_name_raw = row[trib_idx] if trib_idx is not None else None
        skip_vals = ("", "-", "Totale", "TOTALE")
        if not tribunal_name_raw or str(tribunal_name_raw).strip() in skip_vals:
            continue

        tribunal_name = str(tribunal_name_raw).strip()
        district = str(row[col_map["district"]]).strip() if "district" in col_map else ""
        year = _safe_year(row[col_map["year"]])
        if year is None:
            continue

        incoming = _safe_int(row[col_map["incoming"]])
        resolved = _safe_int(row[col_map["resolved"]])
        pending = _safe_int(row[col_map["pending"]])

        records.append(CourtRecord(
            tribunal_name=tribunal_name,
            district=district,
            year=year,
            case_category=case_category,
            incoming=incoming,
            resolved=resolved,
            pending=pending,
        ))

    logger.info("Parsed %d court records from Excel.", len(records))
    return records


# ---------------------------------------------------------------
# Download helper
# ---------------------------------------------------------------


async def download_excel(url: str) -> bytes | None:
    """Download an Excel file. Returns bytes or None on failure."""
    logger.info("Downloading %s ...", url)
    async with httpx.AsyncClient(headers=_HEADERS, timeout=HTTP_TIMEOUT) as client:
        try:
            resp = await client.get(url, follow_redirects=True)
            resp.raise_for_status()
            return resp.content
        except httpx.HTTPStatusError as exc:
            logger.warning("HTTP %d for %s: %s", exc.response.status_code, url, exc)
            return None
        except httpx.RequestError as exc:
            logger.warning("Request error for %s: %s", url, exc)
            return None


# ---------------------------------------------------------------
# Tribunal seed helper
# ---------------------------------------------------------------


async def ensure_tribunals_seeded(session: AsyncSession) -> dict[str, Tribunal]:
    """Ensure all 140+ tribunals exist in the DB. Returns name->Tribunal map.

    Uses a case-insensitive name match. Seeds missing tribunals from the
    static TRIBUNALI list.
    """
    seed_data = get_tribunali()
    seed_by_name: dict[str, TribunalSeed] = {t["name"].lower(): t for t in seed_data}

    # Load existing
    stmt = select(Tribunal)
    result = await session.execute(stmt)
    existing = {t.name.lower(): t for t in result.scalars().all()}

    # Seed missing
    seeded = 0
    for name_lower, seed in seed_by_name.items():
        if name_lower not in existing:
            tribunal = Tribunal(
                name=seed["name"],
                type="ordinario",
                region=seed["region"],
                province=seed["province"],
                lat=seed["lat"],
                lon=seed["lon"],
            )
            session.add(tribunal)
            existing[name_lower] = tribunal
            seeded += 1

    if seeded > 0:
        await session.flush()
        logger.info("Seeded %d new tribunals.", seeded)

    return {name: t for name, t in existing.items()}


def _match_tribunal(
    name: str,
    tribunals_by_name: dict[str, Tribunal],
) -> Tribunal | None:
    """Find a tribunal by name, trying exact match then normalized."""
    key = name.lower().strip()
    if key in tribunals_by_name:
        return tribunals_by_name[key]

    # Try without "Tribunale di" prefix
    for prefix in ("tribunale di ", "tribunale "):
        if key.startswith(prefix):
            stripped = key[len(prefix):]
            if stripped in tribunals_by_name:
                return tribunals_by_name[stripped]

    return None


# ---------------------------------------------------------------
# Ingestor
# ---------------------------------------------------------------


class GiustiziaIngestor(BaseIngestor):
    """Ingest court statistics from Ministero della Giustizia Excel files.

    Source: datiestatistiche.giustizia.it
    Flow:
    1. Download CivileFlussi Excel (or use local file)
    2. Parse rows: tribunal, year, iscritti, definiti, pendenti
    3. Compute clearance_rate and disposition_time
    4. Seed Tribunal table if needed
    5. Upsert CourtStat rows
    """

    source_name = "giustizia"

    async def ingest(
        self,
        session: AsyncSession,
        *,
        limit: int | None = None,
        local_file: str | None = None,
    ) -> int:
        """Fetch and store justice system statistics.

        Args:
            session: Async DB session.
            limit: Max number of records to process (for testing).
            local_file: Path to a local Excel file (skips download).
        """
        log_entry = await self.log_ingestion_start(session)
        total = 0
        errors: dict[str, str] = {}

        try:
            # --- Step 1: Get Excel data ---
            file_content: bytes | None
            if local_file:
                file_content = Path(local_file).read_bytes()
            else:
                file_content = await download_excel(
                    f"{settings.giustizia_stats_url}/resources/CivileFlussi2014-2024.xlsx"
                )
                if file_content is None:
                    # Fallback to known URL
                    file_content = await download_excel(FLUSSI_CIVILE_URL)

                if file_content is None:
                    errors["download"] = "Failed to download justice statistics Excel"
                    await self.log_ingestion_end(
                        session, log_entry, records=0, status="failed", errors=errors,
                    )
                    await session.commit()
                    return 0

            # --- Step 2: Parse Excel ---
            records = parse_excel(file_content, case_category="civile")

            # --- Step 3: Seed tribunals ---
            tribunals_map = await ensure_tribunals_seeded(session)

            # --- Step 4: Upsert CourtStat ---
            unmatched: set[str] = set()
            for i, rec in enumerate(records):
                if limit and i >= limit:
                    break

                tribunal = _match_tribunal(rec.tribunal_name, tribunals_map)
                if tribunal is None:
                    unmatched.add(rec.tribunal_name)
                    continue

                # Upsert: check existing by tribunal_id + period + category
                period = date(rec.year, 12, 31)  # End of year
                stmt = select(CourtStat).where(
                    CourtStat.tribunal_id == tribunal.id,
                    CourtStat.period == period,
                    CourtStat.case_category == rec.case_category,
                )
                result = await session.execute(stmt)
                existing = result.scalar_one_or_none()

                if existing:
                    # Update
                    existing.new_cases = rec.incoming
                    existing.resolved_cases = rec.resolved
                    existing.pending_cases = rec.pending
                    existing.clearance_rate = rec.clearance_rate  # type: ignore[assignment]
                    existing.avg_duration_days = rec.disposition_time  # type: ignore[assignment]
                else:
                    court_stat = CourtStat(
                        tribunal_id=tribunal.id,
                        period=period,
                        case_category=rec.case_category,
                        new_cases=rec.incoming,
                        resolved_cases=rec.resolved,
                        pending_cases=rec.pending,
                        clearance_rate=rec.clearance_rate,
                        avg_duration_days=rec.disposition_time,
                    )
                    session.add(court_stat)
                total += 1

                if total % 500 == 0:
                    await session.flush()
                    logger.info("Giustizia: %d records upserted...", total)

            await session.flush()

            if unmatched:
                logger.warning(
                    "Giustizia: %d tribunal names not matched: %s",
                    len(unmatched),
                    sorted(unmatched)[:10],
                )
                errors["unmatched"] = f"{len(unmatched)} tribunals not matched"

            checkpoint = str(max(r.year for r in records)) if records else None
            await self.log_ingestion_end(
                session, log_entry, records=total, checkpoint=checkpoint,
            )
            await session.commit()
            logger.info("Giustizia: ingestion complete — %d records.", total)

        except Exception as exc:
            errors["exception"] = str(exc)
            await self.log_ingestion_end(
                session, log_entry, records=total, status="failed", errors=errors,
            )
            await session.commit()
            raise

        return total

    async def get_checkpoint(self, session: AsyncSession) -> str | None:
        """Get last ingested year."""
        return await self.get_last_checkpoint(session)
